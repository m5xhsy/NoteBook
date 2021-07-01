from openpyxl import Workbook

wb = Workbook()
# 默认添加到最后
index_sheet = wb.create_sheet("index")
# index_sheet = wb.create_sheet("index",0) # 根据索引添加

###### 添加内容 ######
# 方式一
# index_sheet['A1'] = 4   # 给D3添加aaa
# 方式二
# index_sheet.cell(row=2,column=1,value=6) # 3行5列加value

# 添加行
l = ['姓名', '性别', '年龄']
l2 = ['name','ada','asd']
index_sheet.append(l)
index_sheet.append(l2)
index_sheet['A3'] = "=sum(A1:A2)"   # 公式
index_sheet.title = "inxxx"     # 修改sheet名字

from openpyxl.styles import Font, colors, Alignment
sty = Font(name="等线",size=24,italic=True,color=colors.BLUE,bold=True)
index_sheet['A1'].font = sty

index_sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')


# 第2行行高
index_sheet.row_dimensions[2].height = 40
# C列列宽
index_sheet.column_dimensions['C'].width = 30

# 合并单元格， 往左上角写入数据即可
index_sheet.merge_cells('F1:G1') # 合并一行中的几个单元格
index_sheet.merge_cells('F5:K18') # 合并一个矩形区域中的单元格

# index_sheet.unmerge_cells('F5:K18')
wb.save("G:/index.xlsx")