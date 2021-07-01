from openpyxl import load_workbook
# wb=load_workbook("G:/index.xlsx",read_only=True) # 只读

wb = load_workbook("G:/index.xlsx",data_only=True)  # 将excel的函数转换为值，不写打印出来的是公式

all_sheet = wb.sheetnames
print(all_sheet)

wb_index = wb['index']      # 获取sheet


# 读取值
# 方式一
print(wb_index['A1'].value)
# 方式二
print(wb_index.cell(row=1,column=1).value)

print(wb_index.max_column) # 最大列
print(wb_index.max_row)    # 最大行
print(wb_index.columns)

print([i for i in wb_index.columns]) # 每与列的值
print([i for i in wb_index.rows])    # 每一行的值

for item in wb_index:
    print(item)
    print([val.value for val in item])